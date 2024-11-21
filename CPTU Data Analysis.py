import datetime
import numpy as np
from openpyxl import load_workbook

# ENTERING PROJECT DATA IN FIRST SHEET OF EXCEL
wb = load_workbook('InputOutput.xlsx')
info = wb['Information']
print('**********************************NOTICE*********************************')
print(
    'The author shall not be liable for any direct, consequential or incidental\ndamages arising out of the use of this program. The entire risk as to the \nquality, performance and application of the program lies with the user.')
print('**********************************NOTICE*********************************')
info['C3'].value = input("ENTER THE NAME OF THE PROJECT ")
info['C4'].value = input("ENTER THE PROJECT LOCATION ")
info['C5'].value = input("DESCRIPTION OF THE PROJECT ")
info['C6'].value = str(datetime.datetime.now())

# MANUPLATING RAW CPTU DATA AND DERIVING VARIOUS PARAMATERS
Data = wb['Data']
# SEABED REFERENCING OF TIP RESISTANCE (IN kPA)
for i in range(4, (Data.max_row + 1)):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=5).value is None:
        Data.cell(row=i, column=7).value = None
    else:
        Data.cell(row=i, column=7).value = Data.cell(row=i, column=2).value * 1000 + Data.cell(row=i,
                                                                                               column=5).value * 10 * Data.cell(
            row=1, column=4).value

# SEABED REFERENCING OF POREPRESSURE (IN kPA)
for i in range(4, (Data.max_row + 1)):
    if Data.cell(row=i, column=4).value is None or Data.cell(row=i, column=5).value is None:
        Data.cell(row=i, column=8).value = None
    else:
        Data.cell(row=i, column=8).value = Data.cell(row=i, column=4).value * 1000 + Data.cell(row=i,
                                                                                               column=5).value * 10
# CONE RESISTANCE CORRECTED FOR (U2*) qt (IN kPA)
for i in range(4, (Data.max_row + 1)):
    if Data.cell(row=i, column=7).value is None or Data.cell(row=i, column=8).value is None:
        Data.cell(row=i, column=9).value = None
    else:
        Data.cell(row=i, column=9).value = Data.cell(row=i, column=7).value + Data.cell(row=i, column=8).value * (
                1 - Data.cell(row=1, column=4).value)
# Effective stresses and (0.25 X effective stress)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=6).value is None:
        Data.cell(row=i, column=10).value = None
    else:
        Data.cell(row=i, column=10).value = (Data.cell(row=i, column=6).value * (
                Data.cell(row=i, column=1).value - Data.cell(row=i - 1, column=1).value))
for i in range(4, Data.max_row + 1):
    Data.cell(row=3, column=11).value = 0
    if Data.cell(row=i, column=10).value is None or Data.cell(row=i - 1, column=11).value is None:
        Data.cell(row=i, column=11).value = None
    else:
        Data.cell(row=i, column=11).value = Data.cell(row=i, column=10).value + Data.cell(row=i - 1, column=11).value
    if Data.cell(row=i, column=11).value is None:
        Data.cell(row=i, column=14).value = None
    else:
        Data.cell(row=i, column=14).value = Data.cell(row=i, column=11).value * 0.25
# Total Stresses calculation
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=11).value is None or Data.cell(row=i, column=1).value is None:
        Data.cell(row=i, column=12).value = None
    else:
        Data.cell(row=i, column=12).value = Data.cell(row=i, column=11).value + Data.cell(row=i,
                                                                                          column=1).value * 10  # Normalized cone resistance (Qt) and OCR
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=12).value is None or Data.cell(row=i,
                                                                                                          column=11).value is None:
        Data.cell(row=i, column=13).value = None
    else:
        Data.cell(row=i, column=13).value = (Data.cell(row=i, column=9).value - Data.cell(row=i,
                                                                                          column=12).value) / Data.cell(
            row=i, column=11).value
        Data.cell(row=i, column=31).value = 0.25 * Data.cell(row=i, column=13).value
        Data.cell(row=i, column=32).value = 0.33 * Data.cell(row=i, column=13).value
    if Data.cell(row=i, column=13).value is not None and Data.cell(row=i, column=13).value <= 0:
        Data.cell(row=i, column=13).value = None
        Data.cell(row=i, column=31).value = None
        Data.cell(row=i, column=32).value = None
# Undrained Shear Strength (in kPa) from qc
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=2).value <= 0:
        Data.cell(row=i, column=15).value = None
        Data.cell(row=i, column=16).value = None
    else:
        Data.cell(row=i, column=15).value = Data.cell(row=i, column=2).value * (1000 / Data.cell(row=1, column=8).value)
        Data.cell(row=i, column=16).value = Data.cell(row=i, column=2).value * (
                1000 / Data.cell(row=1, column=10).value)

# Undrained Shear Strength (in kPa) from qc without total stress
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=2).value <= 0 or Data.cell(row=i,
                                                                                                      column=12).value is None:
        Data.cell(row=i, column=17).value = None
        Data.cell(row=i, column=18).value = None
    else:
        Data.cell(row=i, column=17).value = ((Data.cell(row=i, column=2).value * 1000) - Data.cell(row=i,
                                                                                                   column=12).value) / Data.cell(
            row=1, column=8).value
        Data.cell(row=i, column=18).value = ((Data.cell(row=i, column=2).value * 1000) - Data.cell(row=i,
                                                                                                   column=12).value) / Data.cell(
            row=1, column=10).value

# Undrained Shear Strength (in kPa) from qt without total stress
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0 or Data.cell(row=i,
                                                                                                      column=12).value is None:
        Data.cell(row=i, column=19).value = None
        Data.cell(row=i, column=20).value = None
    else:
        Data.cell(row=i, column=19).value = ((Data.cell(row=i, column=9).value) - Data.cell(row=i,
                                                                                            column=12).value) / Data.cell(
            row=1, column=8).value
        Data.cell(row=i, column=20).value = ((Data.cell(row=i, column=9).value) - Data.cell(row=i,
                                                                                            column=12).value) / Data.cell(
            row=1, column=10).value
    if Data.cell(row=i, column=19).value is None or Data.cell(row=i, column=20).value is None or Data.cell(row=i,
                                                                                                           column=19).value <= 0 or Data.cell(
        row=i, column=20).value <= 0:
        Data.cell(row=i, column=19).value = None
        Data.cell(row=i, column=20).value = None

# Undrained Shear Strength (in kPa) from U2
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=8).value is None or Data.cell(row=i, column=1).value is None:
        Data.cell(row=i, column=21).value = None
        Data.cell(row=i, column=22).value = None
    else:
        Data.cell(row=i, column=21).value = (Data.cell(row=i, column=8).value - Data.cell(row=i,
                                                                                          column=1).value * 10) / 6
        Data.cell(row=i, column=22).value = (Data.cell(row=i, column=8).value - Data.cell(row=i,
                                                                                          column=1).value * 10) / 15
    if Data.cell(row=i, column=21).value is None or Data.cell(row=i, column=22).value is None or Data.cell(row=i,
                                                                                                           column=21).value <= 0 or Data.cell(
        row=i, column=22).value <= 0 or Data.cell(row=i, column=21).value is None or Data.cell(row=i,
                                                                                               column=22).value is None:
        Data.cell(row=i, column=21).value = None
        Data.cell(row=i, column=22).value = None

# CONE RESISTANCE  qe (IN kPA)
for i in range(4, (Data.max_row + 1)):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=8).value is None:
        Data.cell(row=i, column=23).value = None
    else:
        Data.cell(row=i, column=23).value = Data.cell(row=i, column=9).value - Data.cell(row=i, column=8).value
# Undrained Shear Strength from qe (in kPa)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=23).value is None or Data.cell(row=i, column=23).value <= 0:
        Data.cell(row=i, column=24).value = None
        Data.cell(row=i, column=25).value = None
    else:
        Data.cell(row=i, column=24).value = Data.cell(row=i, column=23).value / 7
        Data.cell(row=i, column=25).value = Data.cell(row=i, column=23).value / 12
# Friction ratio Rf = (fs/qt)*100
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=9).value is None or Data.cell(row=i,
                                                                                                         column=9).value == 0:
        Data.cell(row=i, column=26).value = None
    else:
        Data.cell(row=i, column=26).value = (Data.cell(row=i, column=3).value * 1000) / Data.cell(row=i,
                                                                                                  column=9).value * 100
    if Data.cell(row=i, column=26).value is None or Data.cell(row=i, column=26).value <= 0:
        Data.cell(row=i, column=26).value = None

# Sensitivity
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=26).value is None or Data.cell(row=i, column=26).value == 0:
        Data.cell(row=i, column=27).value = None
        Data.cell(row=i, column=28).value = None
        Data.cell(row=i, column=29).value = None
    else:
        Data.cell(row=i, column=27).value = 6.0 / Data.cell(row=i, column=26).value
        Data.cell(row=i, column=28).value = 7.5 / Data.cell(row=i, column=26).value
        Data.cell(row=i, column=29).value = 9.0 / Data.cell(row=i, column=26).value
# (u2-u0)/(effective stress) and OCR
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=8).value is None or Data.cell(row=i, column=11).value is None or Data.cell(row=i,
                                                                                                          column=1).value is None:
        Data.cell(row=i, column=30).value = None
    else:
        Data.cell(row=i, column=30).value = (Data.cell(row=i, column=8).value - (
                10 * Data.cell(row=i, column=1).value)) / Data.cell(row=i, column=11).value
        Data.cell(row=i, column=33).value = 0.53 * Data.cell(row=i, column=30).value
    if Data.cell(row=i, column=33).value is not None and Data.cell(row=i, column=33).value <= 0:
        Data.cell(row=i, column=33).value = None  # OCR=0.6*(qt-u2)/effective stress
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=8).value is None or Data.cell(row=i, column=9).value is None or Data.cell(row=i,
                                                                                                         column=11).value is None:
        Data.cell(row=i, column=34).value = None
    else:
        Data.cell(row=i, column=34).value = 0.6 * (
                Data.cell(row=i, column=9).value - Data.cell(row=i, column=8).value) / Data.cell(row=i, column=11).value
    if Data.cell(row=i, column=34).value is not None and Data.cell(row=i, column=34).value <= 0:
        Data.cell(row=i, column=34).value = None
# Fr=fs/(qt-Total stress)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=9).value is None or Data.cell(row=i,
                                                                                                         column=12).value is None:
        Data.cell(row=i, column=35).value = None
    else:
        Data.cell(row=i, column=35).value = 100 * (Data.cell(row=i, column=3).value) / (
                Data.cell(row=i, column=9).value - Data.cell(row=i, column=12).value) * 1000
    if Data.cell(row=i, column=35).value is None or Data.cell(row=i, column=35).value <= 0:
        Data.cell(row=i, column=35).value = None
# Bq=(U2-U0)/(qt-Total stress) and (U2-U0)/U0
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=8).value is None or Data.cell(row=i, column=9).value is None or Data.cell(row=i,
                                                                                                         column=12).value is None or Data.cell(
        row=i, column=1).value is None:
        Data.cell(row=i, column=36).value = None
    else:
        Data.cell(row=i, column=36).value = (Data.cell(row=i, column=8).value - 10 * Data.cell(row=i,
                                                                                               column=1).value) / (
                                                    Data.cell(row=i, column=9).value - Data.cell(row=i,
                                                                                                 column=12).value)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=1).value is None or Data.cell(row=i, column=8).value is None or Data.cell(row=i,
                                                                                                         column=1).value == 0 or Data.cell(
        row=i, column=1).value is None:
        Data.cell(row=i, column=37).value = None
    else:
        Data.cell(row=i, column=37).value = (Data.cell(row=i, column=8).value - 10 * Data.cell(row=i,
                                                                                               column=1).value) / (
                                                    10 * Data.cell(row=i, column=1).value)
# fs/qc
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=3).value is None or Data.cell(row=i,
                                                                                                         column=2).value == 0:
        Data.cell(row=i, column=38).value = None
    else:
        Data.cell(row=i, column=38).value = 100 * Data.cell(row=i, column=3).value / Data.cell(row=i, column=2).value
        Data.cell(row=i, column=41).value = 2 * 1000 * Data.cell(row=i, column=3).value / 3
# U2*/eff.stress
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=8).value is None or Data.cell(row=i, column=11).value is None or Data.cell(row=i,
                                                                                                          column=11).value == 0:
        Data.cell(row=i, column=39).value = None
    else:
        Data.cell(row=i, column=39).value = Data.cell(row=i, column=8).value / Data.cell(row=i, column=11).value
# Sensitivity Robertson 7.1/Fr Robertson 2009
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=35).value is None or Data.cell(row=i, column=35).value == 0:
        Data.cell(row=i, column=40).value = None
    else:
        Data.cell(row=i, column=40).value = 7.1 / Data.cell(row=i, column=35).value
# Su Remoulded
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=3).value == 0:
        Data.cell(row=i, column=41).value = None
    else:
        Data.cell(row=i, column=41).value = 2 * 1000 * Data.cell(row=i, column=3).value / 3
    if Data.cell(row=i, column=41).value is not None and Data.cell(row=i, column=41).value <= 0:
        Data.cell(row=i, column=41).value = None

# (U2-U0)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=1).value is None or Data.cell(row=i, column=8).value is None:
        Data.cell(row=i, column=42).value = None
    else:
        Data.cell(row=i, column=42).value = (Data.cell(row=i, column=8).value - 10 * Data.cell(row=i, column=1).value)
# Relative Density (Dr) Jamilkowski 1988 Ko = 0.5
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=11).value is None or Data.cell(row=i, column=11).value <= 0 or Data.cell(row=i,
                                                                                                        column=2).value is None or Data.cell(
        row=i, column=2).value <= 0:
        Data.cell(row=i, column=43).value = None
    else:
        Data.cell(row=i, column=43).value = (1 / 2.93) * np.log((1000 * Data.cell(row=i, column=2).value / 205) * (
                Data.cell(row=i, column=11).value * (1 + 2 * 0.5) / 3) ** (-0.51))
    if Data.cell(row=i, column=43).value is not None and Data.cell(row=i, column=43).value <= 0:
        Data.cell(row=i, column=43).value = None
# Relative Density (Dr) Jamilkowski 1988 Ko = 1
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=11).value is None or Data.cell(row=i, column=11).value <= 0 or Data.cell(row=i,
                                                                                                        column=2).value is None or Data.cell(
        row=i, column=2).value <= 0:
        Data.cell(row=i, column=44).value = None
    else:
        Data.cell(row=i, column=44).value = (1 / 2.93) * np.log(
            (1000 * Data.cell(row=i, column=2).value / 205) * (Data.cell(row=i, column=11).value * (1 + 2 * 1) / 3) ** (
                -0.51))
    if Data.cell(row=i, column=44).value is not None and Data.cell(row=i, column=44).value <= 0:
        Data.cell(row=i, column=44).value = None
# Relative Density Kulhawy and Mayne (1990)
# Compressibility factor (0.91 for Low, 1.0 for medium and 1.09 for High)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=11).value is None or Data.cell(row=i, column=11).value <= 0 or Data.cell(row=i,
                                                                                                        column=2).value is None or Data.cell(
        row=i, column=2).value <= 0 or Data.cell(row=i, column=45).value is None or Data.cell(row=i,
                                                                                              column=46).value is None:
        Data.cell(row=i, column=47).value = None
    else:
        Data.cell(row=i, column=47).value = (((Data.cell(row=i, column=2).value * 10) / (
                (Data.cell(row=i, column=11).value / 100) ** (0.5))) / (
                                                     305 * Data.cell(row=i, column=45).value * Data.cell(row=i,
                                                                                                         column=46).value * 0.91)) ** (
                                                0.5)
        Data.cell(row=i, column=48).value = (((Data.cell(row=i, column=2).value * 10) / (
                (Data.cell(row=i, column=11).value / 100) ** (0.5))) / (
                                                     305 * Data.cell(row=i, column=45).value * Data.cell(row=i,
                                                                                                         column=46).value * 1.00)) ** (
                                                0.5)
        Data.cell(row=i, column=49).value = (((Data.cell(row=i, column=2).value * 10) / (
                (Data.cell(row=i, column=11).value / 100) ** (0.5))) / (
                                                     305 * Data.cell(row=i, column=45).value * Data.cell(row=i,
                                                                                                         column=46).value * 1.09)) ** (
                                                0.5)
    if Data.cell(row=i, column=47).value is not None and Data.cell(row=i, column=47).value <= 0:
        Data.cell(row=i, column=47).value = None
    if Data.cell(row=i, column=48).value is not None and Data.cell(row=i, column=48).value <= 0:
        Data.cell(row=i, column=48).value = None
    if Data.cell(row=i, column=49).value is not None and Data.cell(row=i, column=49).value <= 0:
        Data.cell(row=i, column=49).value = None
# qt-Total stress
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=12).value is None:
        Data.cell(row=i, column=50).value = None
    else:
        Data.cell(row=i, column=50).value = Data.cell(row=i, column=9).value - Data.cell(row=i, column=12).value
# Phi peak for clays Mayne 2007
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=11).value is None or Data.cell(row=i,
                                                                                                          column=2).value <= 0 or Data.cell(
        row=i, column=11).value <= 0:
        Data.cell(row=i, column=51).value = None
    else:
        Data.cell(row=i, column=51).value = 17.6 + 11 * np.log10(
            (Data.cell(row=i, column=2).value * 10) / ((Data.cell(row=i, column=11).value / 100) ** (0.5)))
# Vs Mayne 2006
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=3).value <= 0:
        Data.cell(row=i, column=52).value = None
    else:
        Data.cell(row=i, column=52).value = 118.8 * np.log10(Data.cell(row=i, column=3).value * 1000) + 18.5
# Vs Long and Donohue (2010)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0 or Data.cell(row=i,
                                                                                                      column=36).value is None or Data.cell(
        row=i, column=36).value <= 0:
        Data.cell(row=i, column=53).value = None
    else:
        Data.cell(row=i, column=53).value = 1.961 * ((Data.cell(row=i, column=9).value) ** (0.579)) * (
                (1 + Data.cell(row=i, column=36).value) ** (1.202))
# Vs Hegazy and Mayne 1995
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0 or Data.cell(row=i,
                                                                                                      column=3).value is None or Data.cell(
        row=i, column=3).value <= 0:
        Data.cell(row=i, column=54).value = None
    else:
        Data.cell(row=i, column=54).value = np.abs(((100 * (Data.cell(row=i, column=3).value * 1000) / Data.cell(row=i,
                                                                                                          column=9).value) ** (
                                                 0.3)) * (10.1 * np.log10(Data.cell(row=i, column=9).value) - 11.4)) ** (
                                                1.67)
# Gmax in MPa

for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0 or Data.cell(row=i,
                                                                                                      column=36).value is None or Data.cell(
        row=i, column=36).value <= 0:
        Data.cell(row=i, column=55).value = None
    else:
        Data.cell(row=i, column=55).value = 4.39 * ((Data.cell(row=i, column=9).value) ** (1.225)) * (
                (1 + Data.cell(row=i, column=36).value) ** (2.53)) / 1000
# SBT Ic
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=13).value is None or Data.cell(row=i, column=13).value <= 0 or Data.cell(row=i,
                                                                                                        column=35).value is None or Data.cell(
        row=i, column=35).value <= 0:
        Data.cell(row=i, column=56).value = None
    else:
        Data.cell(row=i, column=56).value = ((3.47 - np.log10(Data.cell(row=i, column=13).value)) ** (2) + (
                1.22 + np.log10(Data.cell(row=i, column=35).value)) ** (2)) ** (0.5)
    if Data.cell(row=i, column=56).value is None or Data.cell(row=i, column=56).value < 0:
        Data.cell(row=i, column=56).value = None
# Stress Exponent (n)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=56).value is None or Data.cell(row=i, column=11).value is None:
        Data.cell(row=i, column=57).value = None
    else:
        Data.cell(row=i, column=57).value = (
                0.381 * Data.cell(row=i, column=56).value + 0.05 * (Data.cell(row=i, column=11).value / 100) - 0.15)
# Dr Jamilkowski 2001
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=2).value <= 0 or Data.cell(row=i,
                                                                                                      column=11).value is None or Data.cell(
        row=i, column=11).value <= 0:
        Data.cell(row=i, column=58).value = None
    else:
        Data.cell(row=i, column=58).value = 0.268 * np.log(
            (Data.cell(row=i, column=2).value * 10) / ((Data.cell(row=i, column=11).value / 100) ** (0.5))) - 0.675
    if Data.cell(row=i, column=58).value is not None and Data.cell(row=i, column=58).value < 0:
        Data.cell(row=i, column=58).value = None
# Gmax Robertson 2009 
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=50).value is None or Data.cell(row=i, column=50).value <= 0 or Data.cell(row=i,
                                                                                                        column=56).value is None or Data.cell(
        row=i, column=56).value <= 0:
        Data.cell(row=i, column=59).value = None
    else:
        Data.cell(row=i, column=59).value = (Data.cell(row=i, column=50).value / 1000) * (
                0.0188 * 10 ** (0.55 * Data.cell(row=i, column=56).value + 1.68))
    if Data.cell(row=i, column=59).value is None or Data.cell(row=i, column=59).value < 0:
        Data.cell(row=i, column=59).value = None
# Qtn Robertson 2010 
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=50).value is None or Data.cell(row=i, column=50).value <= 0 or Data.cell(row=i,
                                                                                                        column=11).value is None or Data.cell(
        row=i, column=11).value <= 0 or Data.cell(row=i, column=57).value is None or Data.cell(row=i,
                                                                                               column=57).value <= 0:
        Data.cell(row=i, column=60).value = None
    else:
        Data.cell(row=i, column=60).value = (Data.cell(row=i, column=50).value / 100) * (
                (100 / Data.cell(row=i, column=11).value) ** (Data.cell(row=i, column=57).value))
    if Data.cell(row=i, column=60).value is None or Data.cell(row=i, column=60).value < 0:
        Data.cell(row=i, column=60).value = None
# Dimensonless Normalised qc = qc1
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=2).value is None or Data.cell(row=i, column=2).value <= 0 or Data.cell(row=i,
                                                                                                      column=11).value is None or Data.cell(
        row=i, column=11).value <= 0:
        Data.cell(row=i, column=61).value = None
    else:
        Data.cell(row=i, column=61).value = (Data.cell(row=i, column=2).value * 10) / (
                (Data.cell(row=i, column=11).value / 100) ** (0.5))
# Phi'peak (Degrees) General (Mayne)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=36).value is not None and Data.cell(row=i, column=36).value <= 0.1 and Data.cell(row=i,
                                                                                                                column=61).value is not None and Data.cell(
        row=i, column=61).value > 0:
        Data.cell(row=i, column=62).value = 17.6 + 11 * np.log10(Data.cell(row=i, column=61).value)
    elif Data.cell(row=i, column=36).value is not None and Data.cell(row=i, column=36).value >= 0.1 and Data.cell(row=i,
                                                                                                                  column=13).value is not None and Data.cell(
        row=i, column=13).value > 0:
        Data.cell(row=i, column=62).value = 29.5 * ((Data.cell(row=i, column=36).value) ** (0.121)) * (
                0.256 + 0.336 * Data.cell(row=i, column=36).value + np.log10(Data.cell(row=i, column=13).value))
    else:
        Data.cell(row=i, column=62).value = None
# Correction factor for fines, Kc (Robertson, 2010)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=56).value is None or Data.cell(row=i, column=56).value == 0:
        Data.cell(row=i, column=63).value = None
    elif Data.cell(row=i, column=56).value <= 1.64:
        Data.cell(row=i, column=63).value = 1
    else:
        Data.cell(row=i, column=63).value = 5.581 * (Data.cell(row=i, column=56).value) ** (3) - 0.403 * (
            Data.cell(row=i, column=56).value) ** (4) - 21.63 * (Data.cell(row=i, column=56).value) ** (2) + 33.75 * (
                                                Data.cell(row=i, column=56).value) - 17.88
# Qtn, Robertson 2010
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=50).value is None or Data.cell(row=i, column=11).value is None or Data.cell(row=i,
                                                                                                           column=57).value is None:
        Data.cell(row=i, column=64).value = None
    else:
        Data.cell(row=i, column=64).value = (Data.cell(row=i, column=50).value / 100) * (
                100 / Data.cell(row=i, column=11).value) ** (Data.cell(row=i, column=57).value)
    if Data.cell(row=i, column=64).value is None or Data.cell(row=i, column=64).value < 0:
        Data.cell(row=i, column=64).value = None
# Phi, Robertson 2010
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=63).value is None or Data.cell(row=i, column=63).value <= 0 or Data.cell(row=i,
                                                                                                        column=64).value is None or Data.cell(
        row=i, column=64).value <= 0:
        Data.cell(row=i, column=65).value = None
    else:
        Data.cell(row=i, column=65).value = Data.cell(row=1, column=65).value + 15.84 * np.log10(
            Data.cell(row=i, column=63).value * Data.cell(row=i, column=64).value) - 26.88
    if Data.cell(row=i, column=65).value is not None and Data.cell(row=i, column=65).value <= 0:
        Data.cell(row=i, column=65).value = None
# OCR Mayne 2010 and Gmax=p'c*0.24(MPa)
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=56).value is None or Data.cell(row=i, column=56).value == 0:
        Data.cell(row=i, column=66).value = None
        Data.cell(row=i, column=67).value = None
    elif Data.cell(row=i, column=56).value >= 2.95:
        Data.cell(row=i, column=66).value = 0.33 * (Data.cell(row=i, column=50).value) / (
            Data.cell(row=i, column=11).value)
        Data.cell(row=i, column=67).value = 0.24 * Data.cell(row=i, column=11).value * Data.cell(row=i, column=66).value
    else:
        Data.cell(row=i, column=66).value = 0.33 * (Data.cell(row=i, column=50).value) ** (
                0.6 + 0.04 * np.exp(Data.cell(row=i, column=56).value - 0.7)) / (Data.cell(row=i, column=11).value)
        Data.cell(row=i, column=67).value = 0.24 * Data.cell(row=i, column=11).value * Data.cell(row=i, column=66).value
    if Data.cell(row=i, column=66).value is None or Data.cell(row=i, column=66).value < 0:
        Data.cell(row=i, column=66).value = None
    if Data.cell(row=i, column=67).value is None or Data.cell(row=i, column=67).value < 0:
        Data.cell(row=i, column=67).value = None
# Eff.Gamma Mayne 2013
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0 or Data.cell(row=i,
                                                                                                      column=1).value is None or Data.cell(
        row=i, column=1).value == 0:
        Data.cell(row=i, column=68).value = None
    else:
        Data.cell(row=i, column=68).value = 0.056 * (
                (Data.cell(row=i, column=9).value / Data.cell(row=i, column=1).value) ** (1.21))
# Eff.Gamma Mayne 2007 and 2010
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=3).value <= 0 or Data.cell(row=i,
                                                                                                      column=1).value is None or Data.cell(
        row=i, column=9).value is None or Data.cell(row=i, column=9).value <= 0:
        Data.cell(row=i, column=69).value = None
        Data.cell(row=i, column=70).value = None
    else:
        Data.cell(row=i, column=69).value = (2.6 * np.log10(Data.cell(row=i, column=3).value * 1000) + 15 * Data.cell(
            row=1, column=6).value - 26.5) - 10
        Data.cell(row=i, column=70).value = (11.46 + 0.33 * np.log10(Data.cell(row=i, column=1).value) + 3.1 * np.log10(
            Data.cell(row=i, column=3).value * 1000) + 0.7 * np.log10(Data.cell(row=i, column=9).value)) - 10
# Eff.Gamma Robertson 2010
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=26).value is None or Data.cell(row=i, column=26).value <= 0 or Data.cell(row=i,
                                                                                                        column=9).value is None or Data.cell(
        row=i, column=9).value <= 0:
        Data.cell(row=i, column=71).value = None
    else:
        Data.cell(row=i, column=71).value = (0.27 * np.log10(Data.cell(row=i, column=26).value) + 0.36 * np.log10(
            Data.cell(row=i, column=9).value / 100) + 1.236) * (Data.cell(row=1, column=6).value / 2.65) * 10 - 10
# Eff.Gamma Poland
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=3).value <= 0 or Data.cell(row=i,
                                                                                                      column=9).value is None or Data.cell(
        row=i, column=9).value <= 0 or Data.cell(row=i, column=1).value is None:
        Data.cell(row=i, column=72).value = None
    else:
        Data.cell(row=i, column=72).value = (4.077 - 0.5237 * np.log10(
            Data.cell(row=i, column=3).value * 1000) + 5.377 * np.log10(
            Data.cell(row=i, column=9).value) - 2.585 * np.log10(Data.cell(row=i, column=1).value)) - 10
# Eff.Gamma Ghanekar
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=3).value is None or Data.cell(row=i, column=3).value <= 0 or Data.cell(row=i,
                                                                                                      column=9).value is None or Data.cell(
        row=i, column=9).value <= 0 or Data.cell(row=i, column=1).value is None:
        Data.cell(row=i, column=73).value = None
    else:
        Data.cell(row=i, column=73).value = (0.944 - 1.036 * np.log10(
            Data.cell(row=i, column=3).value * 1000) + 6.979 * np.log10(
            Data.cell(row=i, column=9).value) - 2.958 * np.log10(Data.cell(row=i, column=1).value)) - 10
# U0
for i in range(4, Data.max_row + 1):
    if Data.cell(row=i, column=1).value is None:
        Data.cell(row=i, column=74).value = None
    else:
        Data.cell(row=i, column=74).value = Data.cell(row=i, column=1).value * 10


# Saving outputs in excel sheet
wb.save('InputOutput.xlsx')

# Final Completion Message

print('CPTU Data of', info['C5'].value,'Processing is Complete')
